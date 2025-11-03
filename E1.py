import datetime
import sys
from tabulate import tabulate
import sqlite3
from sqlite3 import Error
import os
import csv
import json

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
except Exception:
    openpyxl = None

clientes = []
salas = []
turnos = []
reservas = []

next_cliente_id = 1
next_sala_id = 1
next_folio = 1001

DB_FILE = "Evidencia.db"
FORMATO_FECHA_INPUT = "%m-%d-%Y"   
FORMATO_FECHA_ISO = "%Y-%m-%d"    

def asegurar_tablas():
    crear = False
    if not os.path.exists(DB_FILE):
        crear = True
    else:
        try:
            conexion = sqlite3.connect(DB_FILE)
            cursor = conexion.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='clientes';")
            if not cursor.fetchone():
                crear = True
            cursor.close()
            conexion.close()
        except Exception as err:
            print(f"Error comprobando esquema de BD: {err}")
            crear = True
    
    if crear:
        ddl = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS clientes (
  cliente_id INTEGER PRIMARY KEY AUTOINCREMENT,
  nombre TEXT NOT NULL,
  apellidos TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS salas (
  sala_id INTEGER PRIMARY KEY AUTOINCREMENT,
  nombre TEXT NOT NULL UNIQUE,
  cupo INTEGER NOT NULL CHECK (cupo > 0)
);

CREATE TABLE IF NOT EXISTS turnos (
  turno_id INTEGER PRIMARY KEY,
  descripcion TEXT NOT NULL UNIQUE
);

CREATE TABLE IF NOT EXISTS reservas (
  folio INTEGER PRIMARY KEY AUTOINCREMENT,
  cliente_id INTEGER NOT NULL,
  sala_id INTEGER NOT NULL,
  fecha_normalizada DATE NOT NULL,
  turno_id INTEGER NOT NULL,
  evento TEXT NOT NULL,
  activo INTEGER NOT NULL DEFAULT 1,
  FOREIGN KEY (cliente_id) REFERENCES clientes(cliente_id) ON DELETE CASCADE,
  FOREIGN KEY (sala_id) REFERENCES salas(sala_id) ON DELETE CASCADE,
  FOREIGN KEY (turno_id) REFERENCES turnos(turno_id)
);

-- Eliminar el índice único existente si existe
DROP INDEX IF EXISTS ux_reserva_sala_fecha_turno;

-- Crear índice único parcial que solo aplica a reservas activas
CREATE UNIQUE INDEX IF NOT EXISTS ux_reserva_sala_fecha_turno_activo 
ON reservas (sala_id, fecha_normalizada, turno_id) 
WHERE activo = 1;

CREATE INDEX IF NOT EXISTS ix_reserva_fecha ON reservas (fecha_normalizada);
CREATE INDEX IF NOT EXISTS ix_reserva_cliente ON reservas (cliente_id);

INSERT OR IGNORE INTO turnos (turno_id, descripcion) VALUES (1, 'Matutino');
INSERT OR IGNORE INTO turnos (turno_id, descripcion) VALUES (2, 'Vespertino');
INSERT OR IGNORE INTO turnos (turno_id, descripcion) VALUES (3, 'Nocturno');
"""
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                conexion.executescript(ddl)
        except Error as e:
            print(f"Error al crear tablas en la base de datos: {e}")
            sys.exit(1)
    else:
        # Si la base de datos ya existe, asegurarnos de que el índice único parcial esté creado
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                cursor = conexion.cursor()
                
                # Verificar si el índice único parcial existe
                cursor.execute("SELECT name FROM sqlite_master WHERE type='index' AND name='ux_reserva_sala_fecha_turno_activo'")
                if not cursor.fetchone():
                    # Si no existe, eliminar el índice antiguo y crear el nuevo
                    cursor.execute("DROP INDEX IF EXISTS ux_reserva_sala_fecha_turno")
                    cursor.execute("""
                        CREATE UNIQUE INDEX ux_reserva_sala_fecha_turno_activo 
                        ON reservas (sala_id, fecha_normalizada, turno_id) 
                        WHERE activo = 1
                    """)
                    conexion.commit()
                cursor.close()
        except Exception as err:
            print(f"Error verificando/creando índice único parcial: {err}")

def cargar_estado_desde_bd():
    global clientes, salas, turnos, reservas, next_cliente_id, next_sala_id, next_folio
    asegurar_tablas()
    try:
        with sqlite3.connect(DB_FILE) as conexion:
            conexion.row_factory = sqlite3.Row
            cursor = conexion.cursor()
            
            cursor.execute("SELECT cliente_id AS id, nombre, apellidos FROM clientes ORDER BY apellidos, nombre")
            filas_clientes = cursor.fetchall()
            
            cursor.execute("SELECT sala_id AS id, nombre, cupo FROM salas ORDER BY nombre")
            filas_salas = cursor.fetchall()
            
            cursor.execute("SELECT turno_id, descripcion FROM turnos ORDER BY turno_id")
            filas_turnos = cursor.fetchall()
            
            cursor.execute("""
                SELECT r.folio, r.cliente_id, r.sala_id, r.fecha_normalizada, 
                       t.turno_id, t.descripcion as turno_desc, r.evento, r.activo
                FROM reservas r
                INNER JOIN turnos t ON r.turno_id = t.turno_id
                WHERE r.activo = 1
                ORDER BY r.folio
            """)
            filas_reservas = cursor.fetchall()
            cursor.close()
            
        clientes = [{"id": r["id"], "nombre": r["nombre"], "apellidos": r["apellidos"]} for r in filas_clientes]
        salas = [{"id": r["id"], "nombre": r["nombre"], "cupo": r["cupo"]} for r in filas_salas]
        turnos = [{"id": r["turno_id"], "descripcion": r["descripcion"]} for r in filas_turnos]
        
        reservas = []
        for r in filas_reservas:
            fecha_dt = None
            fecha_texto = r["fecha_normalizada"]
            if fecha_texto:
                try:
                    fecha_dt = datetime.datetime.strptime(fecha_texto, FORMATO_FECHA_ISO).date()
                except Exception:
                    try:
                        fecha_dt = datetime.datetime.strptime(fecha_texto, FORMATO_FECHA_INPUT).date()
                    except Exception:
                        fecha_dt = None
            
            if fecha_dt is None:
                print(f"Advertencia: formato de fecha invalido en BD para folio {r.get('folio')}, registro omitido.")
                continue
                
            reservas.append({
                "folio": r["folio"],
                "cliente_id": r["cliente_id"],
                "sala_id": r["sala_id"],
                "fecha": fecha_dt,
                "turno_id": r["turno_id"],
                "turno": r["turno_desc"],
                "evento": r["evento"],
                "activo": r["activo"]
            })
            
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                cursor = conexion.cursor()
                cursor.execute("SELECT MAX(cliente_id) FROM clientes")
                max_cliente = cursor.fetchone()
                if max_cliente and max_cliente[0]:
                    next_cliente_id = max_cliente[0] + 1
                    
                cursor.execute("SELECT MAX(sala_id) FROM salas")
                max_sala = cursor.fetchone()
                if max_sala and max_sala[0]:
                    next_sala_id = max_sala[0] + 1
                    
                cursor.execute("SELECT MAX(folio) FROM reservas")
                max_folio = cursor.fetchone()
                if max_folio and max_folio[0]:
                    next_folio = max_folio[0] + 1
                    
                cursor.close()
        except Exception as err:
            print(f"Advertencia sincronizando contadores desde BD: {err}")
            
        return True
        
    except Exception as err:
        print(f"No se pudo cargar estado desde BD: {err}")
        return False

def generar_reporte_por_fecha_lista(fecha_consulta):
    filas_reporte = []
    fecha_iso = fecha_consulta.strftime(FORMATO_FECHA_ISO)
    
    try:
        with sqlite3.connect(DB_FILE) as conexion:
            conexion.row_factory = sqlite3.Row
            cursor = conexion.cursor()
            
            query = """
            SELECT 
                r.folio,
                r.fecha_normalizada,
                c.nombre as cliente_nombre,
                c.apellidos as cliente_apellidos,
                s.nombre as sala_nombre,
                s.cupo,
                t.descripcion as turno_descripcion,
                r.evento
            FROM reservas r
            INNER JOIN clientes c ON r.cliente_id = c.cliente_id
            INNER JOIN salas s ON r.sala_id = s.sala_id
            INNER JOIN turnos t ON r.turno_id = t.turno_id
            WHERE r.fecha_normalizada = ? AND r.activo = 1
            ORDER BY r.folio
            """
            
            cursor.execute(query, (fecha_iso,))
            resultados = cursor.fetchall()
            
            for r in resultados:
                filas_reporte.append([
                    r["folio"],
                    fecha_consulta.strftime(FORMATO_FECHA_INPUT),
                    f"{r['cliente_apellidos']}, {r['cliente_nombre']}",
                    r["sala_nombre"],
                    r["cupo"],
                    r["turno_descripcion"],
                    r["evento"]
                ])
            
            cursor.close()
            
    except Exception as err:
        print(f"Error al generar reporte desde BD: {err}")
        for registro_reserva in reservas:
            if registro_reserva["fecha"] == fecha_consulta and registro_reserva["activo"] == 1:
                cliente_encontrado = None
                for registro_cliente in clientes:
                    if registro_cliente["id"] == registro_reserva["cliente_id"]:
                        cliente_encontrado = registro_cliente
                        break
                sala_encontrada = None
                for registro_sala in salas:
                    if registro_sala["id"] == registro_reserva["sala_id"]:
                        sala_encontrada = registro_sala
                        break
                if cliente_encontrado and sala_encontrada:
                    filas_reporte.append([
                        registro_reserva["folio"],
                        registro_reserva["fecha"].strftime(FORMATO_FECHA_INPUT),
                        f"{cliente_encontrado['apellidos']}, {cliente_encontrado['nombre']}",
                        sala_encontrada["nombre"],
                        sala_encontrada["cupo"],
                        registro_reserva["turno"],
                        registro_reserva["evento"]
                    ])
    
    return filas_reporte

def generar_reporte_por_rango_fecha(fecha_inicio, fecha_fin):
    fecha_ini_iso = fecha_inicio.strftime(FORMATO_FECHA_ISO)
    fecha_fin_iso = fecha_fin.strftime(FORMATO_FECHA_ISO)
    
    try:
        with sqlite3.connect(DB_FILE) as conexion:
            conexion.row_factory = sqlite3.Row
            cursor = conexion.cursor()
            
            query = """
            SELECT 
                r.folio,
                r.fecha_normalizada,
                c.nombre as cliente_nombre,
                c.apellidos as cliente_apellidos,
                s.nombre as sala_nombre,
                t.descripcion as turno_descripcion,
                r.evento
            FROM reservas r
            INNER JOIN clientes c ON r.cliente_id = c.cliente_id
            INNER JOIN salas s ON r.sala_id = s.sala_id
            INNER JOIN turnos t ON r.turno_id = t.turno_id
            WHERE r.fecha_normalizada BETWEEN ? AND ? AND r.activo = 1
            ORDER BY r.fecha_normalizada, r.folio
            """
            
            cursor.execute(query, (fecha_ini_iso, fecha_fin_iso))
            resultados = cursor.fetchall()
            cursor.close()
            
            registros = []
            for r in resultados:
                fecha_dt = datetime.datetime.strptime(r["fecha_normalizada"], FORMATO_FECHA_ISO).date()
                registros.append({
                    "folio": r["folio"],
                    "fecha": fecha_dt.strftime(FORMATO_FECHA_INPUT),
                    "cliente": f"{r['cliente_apellidos']}, {r['cliente_nombre']}",
                    "sala": r["sala_nombre"],
                    "turno": r["turno_descripcion"],
                    "evento": r["evento"]
                })
            
            return registros
            
    except Exception as err:
        print(f"Error al obtener reservas por rango: {err}")
        return []

def imprimir_reporte_tabular_por_fecha(fecha_consulta):
    filas = generar_reporte_por_fecha_lista(fecha_consulta)
    if not filas:
        print("\n" + "-" * 60)
        print("NO HAY RESERVACIONES".center(60))
        print("-" * 60)
        print(f"Fecha consultada: {fecha_consulta.strftime(FORMATO_FECHA_INPUT)}")
        print("No se encontraron reservaciones para la fecha indicada.")
        print("-" * 60)
        return False
        
    encabezado = f"REPORTE DE RESERVACIONES PARA EL {fecha_consulta.strftime(FORMATO_FECHA_INPUT)}"
    print("\n" + "=" * 80)
    print(encabezado.center(80))
    print("=" * 80)
    print(tabulate(filas, headers=["FOLIO", "FECHA", "CLIENTE", "SALA", "CUPO", "TURNO", "EVENTO"], tablefmt="grid"))
    print("-" * 80)
    print("FIN DEL REPORTE")
    print("-" * 80)
    return True

def exportar_reporte_json(fecha_consulta, filas_export):
    if not filas_export:
        print("No hay datos para exportar en esa fecha.")
        return
    nombre_archivo = f"reporte_{fecha_consulta.strftime('%Y%m%d')}.json"
    datos_json = []
    for fila in filas_export:
        datos_json.append({
            "folio": fila[0],
            "fecha": fila[1],
            "cliente": fila[2],
            "sala": fila[3],
            "cupo": fila[4],
            "turno": fila[5],
            "evento": fila[6]
        })
    try:
        with open(nombre_archivo, "w", encoding="utf-8") as archivo_salida:
            json.dump(datos_json, archivo_salida, ensure_ascii=False, indent=2)
        print(f"Reporte JSON guardado como: {nombre_archivo}")
    except Exception as err:
        print(f"Error al exportar JSON: {err}")

def exportar_reporte_csv(fecha_consulta, filas_export):
    if not filas_export:
        print("No hay datos para exportar en esa fecha.")
        return
    nombre_archivo = f"reporte_{fecha_consulta.strftime('%Y%m%d')}.csv"
    try:
        with open(nombre_archivo, "w", newline='', encoding="utf-8") as archivo_csv:
            escritor = csv.writer(archivo_csv)
            escritor.writerow(["FOLIO","FECHA","CLIENTE","SALA","CUPO","TURNO","EVENTO"])
            for fila in filas_export:
                escritor.writerow(fila)
        print(f"Reporte CSV guardado como: {nombre_archivo}")
    except Exception as err:
        print(f"Error al exportar CSV: {err}")

def exportar_reporte_excel(fecha_consulta, filas_export):
    if openpyxl is None:
        print("openpyxl no esta instalado. Instale openpyxl para exportar a Excel.")
        return
    if not filas_export:
        print("No hay datos para exportar en esa fecha.")
        return
    nombre_archivo = f"reporte_{fecha_consulta.strftime('%Y%m%d')}.xlsx"
    try:
        libro = openpyxl.Workbook()
        hoja = libro.active
        hoja.title = "Reservas"
        hoja.merge_cells('A1:G1')
        hoja['A1'] = f"REPORTE RESERVACIONES {fecha_consulta.strftime(FORMATO_FECHA_INPUT)}"
        hoja['A1'].font = Font(bold=True)
        encabezados = ["FOLIO","FECHA","CLIENTE","SALA","CUPO","TURNO","EVENTO"]
        for indice_columna, titulo_columna in enumerate(encabezados, start=1):
            celda = hoja.cell(row=2, column=indice_columna, value=titulo_columna)
            celda.font = Font(bold=True)
            celda.alignment = Alignment(horizontal="center")
            celda.border = Border(bottom=Side(border_style="thick"))
        for indice_fila, fila_datos in enumerate(filas_export, start=3):
            for indice_columna, valor in enumerate(fila_datos, start=1):
                celda = hoja.cell(row=indice_fila, column=indice_columna, value=valor)
                celda.alignment = Alignment(horizontal="center")
        libro.save(nombre_archivo)
        print(f"Reporte Excel guardado como: {nombre_archivo}")
    except Exception as err:
        print(f"Error al exportar Excel: {err}")

# Inicialización
inicio_bd_ok = cargar_estado_desde_bd()
if inicio_bd_ok:
    print("\n" + "=" * 70)
    print("Estado inicial cargado desde Evidencia.db".center(70))
    print("=" * 70)
else:
    print("\n" + "=" * 70)
    print("No se pudo cargar Evidencia.db; iniciando con estado vacio".center(70))
    print("=" * 70)

# Menú principal
while True:
    print("\n" + "=" * 60)
    print("SISTEMA DE RESERVACION DE SALAS".center(60))
    print("=" * 60)
    print("1. Registrar reservacion de una sala.")
    print("2. Cancelar evento.")
    print("3. Editar nombre de evento.")
    print("4. Consultar reservaciones por fecha.")
    print("5. Registrar un nuevo cliente.")
    print("6. Registrar una sala.")
    print("7. Salir.")
    print("=" * 60)

    try:
        opcion_texto = input("Seleccionar una opcion (1-7): ").strip()
    except (EOFError, KeyboardInterrupt):
        print("\nOperacion cancelada por el usuario.")
        sys.exit()
        
    if opcion_texto == "":
        print("Entrada vacia: ingrese un numero entre 1 and 7.")
        continue
    if not opcion_texto.isdigit():
        print("Formato invalido: la opcion debe ser numerica entre 1 and 7.")
        continue
        
    opcion = int(opcion_texto)
    if opcion < 1 or opcion > 7:
        print("Opcion fuera de rango: seleccione un valor entre 1 and 7.")
        continue

    if opcion == 1:
        print("\n" + "=" * 60)
        print("REGISTRAR RESERVACION".center(60))
        print("=" * 60)
        cancelar = False

        # CORRECCIÓN APLICADA: Manejo correcto de fechas domingo
        # Ingreso de fecha
        while True:
            try:
                texto_fecha = input("\nIngrese fecha de reservacion (MM-DD-YYYY) o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar = True
                break
                
            if texto_fecha.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar = True
                break
                
            if texto_fecha == "":
                print("Fecha invalida: el campo 'Fecha' esta vacio. Formato esperado: MM-DD-YYYY.")
                continue
                
            # Validación de fecha
            if any(c.isalpha() for c in texto_fecha):
                print("Fecha invalida: hay letras en la fecha. Use solo digitos y guiones, ejemplo: 12-31-2025.")
                continue
            if any(c in ",./\\" for c in texto_fecha) and "-" not in texto_fecha:
                print("Fecha invalida: separadores incorrectos. Use '-' entre mes, dia y año. Ejemplo: MM-DD-YYYY.")
                continue
            try:
                fecha = datetime.datetime.strptime(texto_fecha, FORMATO_FECHA_INPUT).date()
            except ValueError:
                print("Fecha invalida: formato incorrecto. Use MM-DD-YYYY, ejemplo: 12-31-2025.")
                continue
                
            if fecha < datetime.date.today() + datetime.timedelta(days=2):
                print("Restriccion de antelacion: la fecha debe ser al menos dos dias posterior a hoy.")
                continue
                
            if fecha.weekday() == 6:
                lunes_propuesto = fecha + datetime.timedelta(days=1)
                while True:
                    try:
                        respuesta_domingo = input(f"La fecha ingresada es domingo. Se propone {lunes_propuesto.strftime(FORMATO_FECHA_INPUT)}. Aceptar? (S/N) o 'X' para cancelar: ").strip().upper()
                    except (EOFError, KeyboardInterrupt):
                        print("\nOperacion cancelada por el usuario.")
                        respuesta_domingo = "X"
                        
                    if respuesta_domingo == "X":
                        cancelar = True
                        break
                    if respuesta_domingo == "":
                        print("Respuesta vacia: escriba 'S' para aceptar o 'N' para rechazar.")
                        continue
                    if respuesta_domingo not in ("S", "N"):
                        print("Respuesta invalida: escriba 'S' para aceptar o 'N' para rechazar.")
                        continue
                    if respuesta_domingo == "S":
                        fecha = lunes_propuesto
                        break
                    else:  # Respuesta "N"
                        print("Fecha domingo rechazada. Por favor ingrese una nueva fecha que no sea domingo.")
                        break  # Rompe el bucle interno para volver a pedir fecha
                        
                if cancelar:
                    break
                    
                if respuesta_domingo == "N":
                    continue  # Vuelve al inicio del bucle principal para pedir nueva fecha
                else:
                    break  # Fecha aceptada (lunes) o cancelada
                    
            else:  # No es domingo
                print(f"Fecha aceptada: {fecha.strftime(FORMATO_FECHA_INPUT)}")
                break
                
        if cancelar:
            continue

        # Selección de cliente
        cliente_nombre_completo = ""
        while True:
            clientes_bd = []
            try:
                with sqlite3.connect(DB_FILE) as conexion:
                    conexion.row_factory = sqlite3.Row
                    cursor = conexion.cursor()
                    cursor.execute("SELECT cliente_id, apellidos, nombre FROM clientes ORDER BY apellidos, nombre")
                    clientes_bd = cursor.fetchall()
                    cursor.close()
            except Exception as err:
                print(f"Error al leer lista de clientes desde BD: {err}")
                clientes_bd = []

            if not clientes_bd:
                print("\nNo hay clientes registrados. Use la opcion 5 para registrar un cliente.")
                cancelar = True
                break

            print("\n" + "-" * 50)
            print("CLIENTES REGISTRADOS")
            print("-" * 50)
            for fila_cliente in clientes_bd:
                print(f"{fila_cliente['cliente_id']}: {fila_cliente['apellidos']}, {fila_cliente['nombre']}")
                
            try:
                sel_cliente_texto = input("\nIngrese ID de cliente o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar = True
                break
                
            if sel_cliente_texto.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar = True
                break
                
            if sel_cliente_texto == "":
                print("ID invalido: el campo esta vacio.")
                continue
            if sel_cliente_texto == "0":
                print("ID invalido: el numero debe ser mayor a 0.")
                continue
            if not sel_cliente_texto.isdigit():
                print("ID invalido: no se aceptan letras en el ID del cliente.")
                continue
                
            cliente_id = int(sel_cliente_texto)
            encontrado = any(fila['cliente_id'] == cliente_id for fila in clientes_bd)
            if not encontrado:
                print(f"ID {cliente_id} no encontrado en la base de datos. Ingrese un ID valido de la lista.")
                continue
                
            cliente_seleccionado = next((fila for fila in clientes_bd if fila['cliente_id'] == cliente_id), None)
            if cliente_seleccionado:
                cliente_nombre_completo = f"{cliente_seleccionado['apellidos']}, {cliente_seleccionado['nombre']}"
                print(f"Cliente seleccionado: {cliente_nombre_completo}")
            break
            
        if cancelar:
            continue

        # Verificar disponibilidad
        fecha_norm_texto = fecha.strftime(FORMATO_FECHA_ISO)
        disponibles = []
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                conexion.row_factory = sqlite3.Row
                cursor = conexion.cursor()
                
                cursor.execute("SELECT sala_id, nombre, cupo FROM salas ORDER BY nombre")
                filas_salas = cursor.fetchall()
                
                cursor.execute("SELECT turno_id, descripcion FROM turnos ORDER BY turno_id")
                filas_turnos = cursor.fetchall()
                cursor.close()
                
            for fila_sala in filas_salas:
                sala_id_tmp = fila_sala['sala_id']
                sala_nombre_tmp = fila_sala['nombre']
                sala_cupo_tmp = fila_sala['cupo']
                
                for fila_turno in filas_turnos:
                    turno_id_tmp = fila_turno['turno_id']
                    turno_desc_tmp = fila_turno['descripcion']
                    
                    try:
                        with sqlite3.connect(DB_FILE) as conexion_check:
                            cursor_check = conexion_check.cursor()
                            cursor_check.execute("SELECT 1 FROM reservas WHERE sala_id=? AND fecha_normalizada=? AND turno_id=? AND activo=1",
                                                 (sala_id_tmp, fecha_norm_texto, turno_id_tmp))
                            ocupado_bd = cursor_check.fetchone() is not None
                            cursor_check.close()
                    except Exception as err:
                        print(f"Error verificando disponibilidad: {err}")
                        ocupado_bd = True
                        
                    if not ocupado_bd:
                        disponibles.append((sala_id_tmp, sala_nombre_tmp, sala_cupo_tmp, turno_desc_tmp))
                        
        except Exception as err:
            print(f"Error al leer salas desde BD: {err}")
            disponibles = []

        if not disponibles:
            print("\n" + "-" * 60)
            print("NO HAY SALAS DISPONIBLES")
            print("-" * 60)
            print(f"Para la fecha: {fecha.strftime(FORMATO_FECHA_INPUT)}")
            print("No existen salas con turnos libres.")
            print("\nSugerencias:")
            print("Seleccione otra fecha")
            print("Registre mas salas (Opcion 6)")
            print("-" * 60)
            continue

        # Mostrar salas disponibles
        print("\n" + "-" * 50)
        print(f"SALAS DISPONIBLES PARA {fecha.strftime(FORMATO_FECHA_INPUT)}")
        print("-" * 50)
        salas_mostradas = set()
        for registro_disponible in disponibles:
            id_sala_disp = registro_disponible[0]
            nombre_sala_disp = registro_disponible[1]
            cupo_sala_disp = registro_disponible[2]
            
            if id_sala_disp not in salas_mostradas:
                print(f"\nSALA {id_sala_disp}: {nombre_sala_disp} (Cupo: {cupo_sala_disp} personas)")
                salas_mostradas.add(id_sala_disp)
            
            turno_disp = registro_disponible[3]
            print(f"   {turno_disp}")

        # Selección de sala
        sala_nombre = ""
        while True:
            try:
                sel_sala_texto = input("\nIngrese ID de sala o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar = True
                break
                
            if sel_sala_texto.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar = True
                break
                
            if sel_sala_texto == "":
                print("ID de sala invalido: campo vacio.")
                continue
            if sel_sala_texto == "0":
                print("ID invalido: el numero debe ser mayor a 0.")
                continue
            if not sel_sala_texto.isdigit():
                print("ID de sala invalido: no se aceptan letras en el ID.")
                continue
                
            sala_id = int(sel_sala_texto)

            try:
                with sqlite3.connect(DB_FILE) as conexion:
                    cursor = conexion.cursor()
                    cursor.execute("SELECT sala_id, nombre FROM salas WHERE sala_id=?", (sala_id,))
                    resultado = cursor.fetchone()
                    cursor.close()
                    
                if not resultado:
                    print(f"ID {sala_id} no encontrado en la base de datos. Ingrese un ID valido.")
                    continue
                else:
                    sala_nombre = resultado[1]
                    
            except Exception as err:
                print(f"Error verificando sala: {err}")
                continue

            existe_en_lista = any(registro_disponible[0] == sala_id for registro_disponible in disponibles)
            if not existe_en_lista:
                print(f"La sala {sala_id} no tiene turnos disponibles para esta fecha.")
                print("Seleccione otra sala de la lista.")
                continue
                
            break

        if cancelar:
            continue

        # Selección de turno
        lista_turnos_disponibles = []
        for registro_disponible in disponibles:
            if registro_disponible[0] == sala_id:
                lista_turnos_disponibles.append(registro_disponible[3])

        print("\nSELECCIONE EL TURNO")
        for indice, descripcion_turno in enumerate(("Matutino", "Vespertino", "Nocturno"), start=1):
            disponible_texto = "DISPONIBLE" if descripcion_turno in lista_turnos_disponibles else "OCUPADO"
            print(f"{indice}. {descripcion_turno} - {disponible_texto}")
        print("X. Cancelar operacion")
        
        while True:
            try:
                sel_turno_texto = input("\nElija el numero de turno (1-3) o 'X': ").strip().upper()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar = True
                break
                
            if sel_turno_texto == "X":
                print("Operacion cancelada por el usuario.")
                cancelar = True
                break
                
            if sel_turno_texto == "":
                print("Seleccion invalida: campo vacio.")
                continue
            if not sel_turno_texto.isdigit():
                print("Seleccion invalida: no se aceptan letras para seleccionar turno.")
                continue
                
            num_turno = int(sel_turno_texto)
            if num_turno not in (1, 2, 3):
                print("Seleccion fuera de rango: elija 1, 2 o 3.")
                continue
                
            turno_seleccionado = {1: "Matutino", 2: "Vespertino", 3: "Nocturno"}[num_turno]
            if turno_seleccionado not in lista_turnos_disponibles:
                print(f"Turno {turno_seleccionado} no disponible para la sala seleccionada.")
                print("Elija otro turno disponible.")
                continue
                
            break
            
        if cancelar:
            continue

        # Ingreso del nombre del evento
        while True:
            try:
                nombre_evento_texto = input("\nNombre del evento o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar = True
                break
                
            if nombre_evento_texto.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar = True
                break
                
            # Validación de nombre de evento
            texto_limpio = nombre_evento_texto.strip()
            if not texto_limpio:
                print("El nombre del evento no puede estar vacio")
                print("Ejemplo valido: 'Reunion de equipo', 'Conferencia anual', 'Waifu Fest 2'")
                continue
            if len(texto_limpio) < 3:
                print("El nombre del evento debe tener al menos 3 caracteres")
                print("Ejemplo valido: 'Reunion de equipo', 'Conferencia anual', 'Waifu Fest 2'")
                continue
            if all(c in ' \t\n' for c in texto_limpio):
                print("El nombre del evento no puede contener solo espacios")
                print("Ejemplo valido: 'Reunion de equipo', 'Conferencia anual', 'Waifu Fest 2'")
                continue
                
            break
            
        if cancelar:
            continue

        # Insertar reserva
        fecha_norm_texto = fecha.strftime(FORMATO_FECHA_ISO)
        
        # Obtener turno_id
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                cursor = conexion.cursor()
                cursor.execute("SELECT turno_id FROM turnos WHERE descripcion = ?", (turno_seleccionado,))
                resultado = cursor.fetchone()
                cursor.close()
                turno_id = resultado[0] if resultado else None
        except Exception as err:
            print(f"Error al obtener ID del turno: {err}")
            turno_id = None

        if not turno_id:
            print(f"Error: Turno '{turno_seleccionado}' no encontrado")
            continue

        # Verificar conflicto e insertar
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                conexion.row_factory = sqlite3.Row
                cursor = conexion.cursor()

                cursor.execute("SELECT 1 FROM reservas WHERE sala_id=? AND fecha_normalizada=? AND turno_id=? AND activo=1", 
                              (sala_id, fecha_norm_texto, turno_id))
                if cursor.fetchone():
                    print("Error: Ya existe una reserva activa para esa sala, fecha y turno")
                    continue

                cursor.execute("INSERT INTO reservas (cliente_id, sala_id, fecha_normalizada, turno_id, evento) VALUES (?,?,?,?,?)",
                              (cliente_id, sala_id, fecha_norm_texto, turno_id, nombre_evento_texto))
                conexion.commit()
                folio_generado = cursor.lastrowid
                cursor.close()
                
            cargar_estado_desde_bd()
            print("\n" + "=" * 60)
            print("RESERVACION REGISTRADA EXITOSAMENTE")
            print("=" * 60)
            print(f"Folio: {folio_generado}")
            print(f"Cliente: {cliente_nombre_completo}")
            print(f"Sala: {sala_nombre}")
            print(f"Fecha: {fecha.strftime(FORMATO_FECHA_INPUT)}")
            print(f"Turno: {turno_seleccionado}")
            print(f"Evento: {nombre_evento_texto}")
            print("=" * 60)
            
        except sqlite3.IntegrityError as err:
            print(f"Reserva no insertada en BD (error de integridad): {err}")
            print("Nota: Esto puede ocurrir si hay un conflicto de unicidad. Verifique que no exista una reserva activa para la misma sala, fecha y turno.")
        except Exception as err:
            print(f"Error al insertar reserva en BD: {err}")

    elif opcion == 2:
        print("\n" + "=" * 60)
        print("CANCELAR RESERVACION")
        print("=" * 60)
        cancelar_operacion = False

        # Solicitar rango de fechas
        while True:
            try:
                texto_fecha_ini = input("\nFecha inicial (MM-DD-YYYY) o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_ini.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_ini == "":
                print("Fecha inicial invalida: campo vacio.")
                continue
                
            # Validación de fecha
            if any(c.isalpha() for c in texto_fecha_ini):
                print("Fecha inicial invalida. Use formato MM-DD-YYYY.")
                continue
            if any(c in ",./\\" for c in texto_fecha_ini) and "-" not in texto_fecha_ini:
                print("Fecha inicial invalida. Use formato MM-DD-YYYY.")
                continue
            try:
                fecha_inicio = datetime.datetime.strptime(texto_fecha_ini, FORMATO_FECHA_INPUT).date()
            except ValueError:
                print("Fecha inicial invalida. Use formato MM-DD-YYYY.")
                continue
                
            break
            
        if cancelar_operacion:
            continue

        while True:
            try:
                texto_fecha_fin = input("Fecha final (MM-DD-YYYY) o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_fin.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_fin == "":
                print("Fecha final invalida: campo vacio.")
                continue
                
            # Validación de fecha
            if any(c.isalpha() for c in texto_fecha_fin):
                print("Fecha final invalida. Use formato MM-DD-YYYY.")
                continue
            if any(c in ",./\\" for c in texto_fecha_fin) and "-" not in texto_fecha_fin:
                print("Fecha final invalida. Use formato MM-DD-YYYY.")
                continue
            try:
                fecha_fin = datetime.datetime.strptime(texto_fecha_fin, FORMATO_FECHA_INPUT).date()
            except ValueError:
                print("Fecha final invalida. Use formato MM-DD-YYYY.")
                continue
                
            break
            
        if cancelar_operacion:
            continue

        if fecha_fin < fecha_inicio:
            print("Rango invalido: la fecha final es anterior a la inicial.")
            continue

        # Obtener reservas en el rango usando la función
        reservas_rango = generar_reporte_por_rango_fecha(fecha_inicio, fecha_fin)
        
        if not reservas_rango:
            print(f"\nNo hay reservaciones activas entre {fecha_inicio.strftime(FORMATO_FECHA_INPUT)} y {fecha_fin.strftime(FORMATO_FECHA_INPUT)}")
            continue

        # Mostrar reservas
        print("\n" + "-" * 50)
        print(f"RESERVACIONES DEL {fecha_inicio.strftime(FORMATO_FECHA_INPUT)} AL {fecha_fin.strftime(FORMATO_FECHA_INPUT)}")
        print("-" * 50)
        tabla_reservas = [[r["folio"], r["fecha"], r["cliente"], r["sala"], r["turno"], r["evento"]] for r in reservas_rango]
        print(tabulate(tabla_reservas, headers=["FOLIO", "FECHA", "CLIENTE", "SALA", "TURNO", "EVENTO"], tablefmt="grid"))

        # Seleccionar folio a cancelar
        while True:
            try:
                folio_cancelar_texto = input("\nIngrese el folio a cancelar o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if folio_cancelar_texto.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if folio_cancelar_texto == "":
                print("Folio invalido: campo vacio.")
                continue
            if not folio_cancelar_texto.isdigit():
                print("Folio invalido: debe ser un numero.")
                continue
                
            folio_cancelar = int(folio_cancelar_texto)
            
            # Verificar que el folio existe en el rango
            reserva_encontrada = next((r for r in reservas_rango if r["folio"] == folio_cancelar), None)
            if not reserva_encontrada:
                print(f"Folio {folio_cancelar} no encontrado en el rango especificado.")
                continue
                
            # Verificar anticipación (mínimo 2 días)
            fecha_reserva = datetime.datetime.strptime(reserva_encontrada["fecha"], FORMATO_FECHA_INPUT).date()
            dias_restantes = (fecha_reserva - datetime.date.today()).days
            
            if dias_restantes < 2:
                print(f"No se puede cancelar: faltan {dias_restantes} dia(s).")
                print("Se requiere al menos 2 dias de anticipacion para cancelar.")
                break
                
            # Confirmar cancelación
            try:
                confirmacion = input(f"Esta seguro de cancelar la reservacion folio {folio_cancelar}? (S/N): ").strip().upper()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if confirmacion != "S":
                print("Cancelacion abortada por el usuario.")
                break
                
            # Ejecutar cancelación (actualizar activo = 0)
            try:
                with sqlite3.connect(DB_FILE) as conexion:
                    cursor = conexion.cursor()
                    cursor.execute("UPDATE reservas SET activo = 0 WHERE folio = ?", (folio_cancelar,))
                    conexion.commit()
                    cursor.close()
                cargar_estado_desde_bd()
                print(f"Reservacion folio {folio_cancelar} cancelada exitosamente.")
                print("La reserva ya no aparecera en los reportes del sistema.")
            except Exception as err:
                print(f"Error al cancelar la reservacion folio {folio_cancelar}: {err}")
                
            break
            
        if cancelar_operacion:
            continue

    elif opcion == 3:
        print("\n" + "=" * 60)
        print("EDITAR NOMBRE DE EVENTO")
        print("=" * 60)
        cancelar_operacion = False

        # Solicitar rango de fechas
        while True:
            try:
                texto_fecha_ini = input("\nFecha inicial (MM-DD-YYYY) o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_ini.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_ini == "":
                print("Fecha inicial invalida: campo vacio.")
                continue
                
            # Validación de fecha
            if any(c.isalpha() for c in texto_fecha_ini):
                print("Fecha inicial invalida. Use formato MM-DD-YYYY.")
                continue
            if any(c in ",./\\" for c in texto_fecha_ini) and "-" not in texto_fecha_ini:
                print("Fecha inicial invalida. Use formato MM-DD-YYYY.")
                continue
            try:
                fecha_inicio = datetime.datetime.strptime(texto_fecha_ini, FORMATO_FECHA_INPUT).date()
            except ValueError:
                print("Fecha inicial invalida. Use formato MM-DD-YYYY.")
                continue
                
            break
            
        if cancelar_operacion:
            continue

        while True:
            try:
                texto_fecha_fin = input("Fecha final (MM-DD-YYYY) o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_fin.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if texto_fecha_fin == "":
                print("Fecha final invalida: campo vacio.")
                continue
                
            # Validación de fecha
            if any(c.isalpha() for c in texto_fecha_fin):
                print("Fecha final invalida. Use formato MM-DD-YYYY.")
                continue
            if any(c in ",./\\" for c in texto_fecha_fin) and "-" not in texto_fecha_fin:
                print("Fecha final invalida. Use formato MM-DD-YYYY.")
                continue
            try:
                fecha_fin = datetime.datetime.strptime(texto_fecha_fin, FORMATO_FECHA_INPUT).date()
            except ValueError:
                print("Fecha final invalida. Use formato MM-DD-YYYY.")
                continue
                
            break
            
        if cancelar_operacion:
            continue

        if fecha_fin < fecha_inicio:
            print("Rango invalido: la fecha final es anterior a la inicial.")
            continue

        # Obtener reservas en el rango usando la función
        reservas_rango = generar_reporte_por_rango_fecha(fecha_inicio, fecha_fin)
        
        if not reservas_rango:
            print(f"\nNo hay reservaciones activas entre {fecha_inicio.strftime(FORMATO_FECHA_INPUT)} y {fecha_fin.strftime(FORMATO_FECHA_INPUT)}")
            continue

        # Mostrar reservas
        print("\n" + "-" * 50)
        print(f"RESERVACIONES DEL {fecha_inicio.strftime(FORMATO_FECHA_INPUT)} AL {fecha_fin.strftime(FORMATO_FECHA_INPUT)}")
        print("-" * 50)
        tabla_reservas = [[r["folio"], r["fecha"], r["cliente"], r["sala"], r["turno"], r["evento"]] for r in reservas_rango]
        print(tabulate(tabla_reservas, headers=["FOLIO", "FECHA", "CLIENTE", "SALA", "TURNO", "EVENTO"], tablefmt="grid"))

        # Seleccionar folio a editar
        while True:
            try:
                folio_editar_texto = input("\nIngrese el folio a editar o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if folio_editar_texto.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if folio_editar_texto == "":
                print("Folio invalido: campo vacio.")
                continue
            if not folio_editar_texto.isdigit():
                print("Folio invalido: debe ser un numero.")
                continue
                
            folio_editar = int(folio_editar_texto)
            
            # Verificar que el folio existe en el rango
            reserva_encontrada = next((r for r in reservas_rango if r["folio"] == folio_editar), None)
            if not reserva_encontrada:
                print(f"Folio {folio_editar} no encontrado en el rango especificado.")
                continue
                
            break
            
        if cancelar_operacion:
            continue

        # Solicitar nuevo nombre del evento
        while True:
            try:
                nuevo_nombre = input("\nNuevo nombre del evento o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            if nuevo_nombre.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_operacion = True
                break
                
            # Validación de nombre de evento
            texto_limpio = nuevo_nombre.strip()
            if not texto_limpio:
                print("El nombre del evento no puede estar vacio")
                print("Ejemplo valido: 'Reunion de equipo', 'Conferencia anual', 'Waifu Fest 2'")
                continue
            if len(texto_limpio) < 3:
                print("El nombre del evento debe tener al menos 3 caracteres")
                print("Ejemplo valido: 'Reunion de equipo', 'Conferencia anual', 'Waifu Fest 2'")
                continue
            if all(c in ' \t\n' for c in texto_limpio):
                print("El nombre del evento no puede contener solo espacios")
                print("Ejemplo valido: 'Reunion de equipo', 'Conferencia anual', 'Waifu Fest 2'")
                continue
                
            break
            
        if cancelar_operacion:
            continue

        # Confirmar edición
        try:
            confirmacion = input(f"Esta seguro de cambiar el nombre del evento folio {folio_editar}? (S/N): ").strip().upper()
        except (EOFError, KeyboardInterrupt):
            print("\nOperacion cancelada por el usuario.")
            continue
            
        if confirmacion != "S":
            print("Edicion abortada por el usuario.")
            continue

        # Ejecutar actualización
        try:
            with sqlite3.connect(DB_FILE) as conexion:
                cursor = conexion.cursor()
                cursor.execute("UPDATE reservas SET evento = ? WHERE folio = ?", (nuevo_nombre, folio_editar))
                conexion.commit()
                cursor.close()
            cargar_estado_desde_bd()
            print(f"Evento folio {folio_editar} actualizado exitosamente.")
            print(f"Nuevo nombre: {nuevo_nombre}")
        except Exception as err:
            print(f"Error al actualizar el evento folio {folio_editar}: {err}")

    elif opcion == 4:
        print("\n" + "=" * 60)
        print("CONSULTAR RESERVACIONES POR FECHA")
        print("=" * 60)

        while True:
            try:
                texto_fecha_consulta = input("\nIngrese la fecha a consultar (MM-DD-YYYY) o Enter para hoy: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                break

            if texto_fecha_consulta == "":
                fecha_consulta = datetime.date.today()
                print(f"\nFecha consultada: {fecha_consulta.strftime(FORMATO_FECHA_INPUT)} (hoy)")
            else:
                # Validación de fecha
                if any(c.isalpha() for c in texto_fecha_consulta):
                    print("Fecha invalida: hay letras en la fecha. Use solo digitos y guiones.")
                    continue
                if any(c in ",./\\" for c in texto_fecha_consulta) and "-" not in texto_fecha_consulta:
                    print("Fecha invalida: separadores incorrectos. Use '-' entre mes, dia y año.")
                    continue
                try:
                    fecha_consulta = datetime.datetime.strptime(texto_fecha_consulta, FORMATO_FECHA_INPUT).date()
                except ValueError:
                    print("Fecha invalida: formato incorrecto. Use MM-DD-YYYY.")
                    continue
                print(f"\nFecha consultada: {fecha_consulta.strftime(FORMATO_FECHA_INPUT)}")

            hay_registros = imprimir_reporte_tabular_por_fecha(fecha_consulta)
            
            if not hay_registros:
                while True:
                    try:
                        resp_no_reg = input("\nDesea consultar otra fecha? (S/N): ").strip().upper()
                    except (EOFError, KeyboardInterrupt):
                        print("\nOperacion cancelada por el usuario.")
                        resp_no_reg = "N"
                        break
                        
                    if resp_no_reg == "":
                        print("Respuesta vacia: escriba 'S' para si o 'N' para no.")
                        continue
                    if resp_no_reg not in ("S", "N"):
                        print("Respuesta invalida: escriba 'S' para si o 'N' para no.")
                        continue
                        
                    if resp_no_reg == "S":
                        break
                    else:
                        break
                if resp_no_reg == "N":
                    break
                else:
                    continue

            # Opciones de exportación
            print("\n" + "-" * 50)
            print("OPCIONES DE EXPORTACION")
            print("-" * 50)
            print("a) Exportar a CSV")
            print("b) Exportar a JSON") 
            print("c) Exportar a Excel")
            print("d) No exportar (regresar al menu)")
            
            while True:
                try:
                    opcion_export_texto = input("\nSeleccione una opcion (a/b/c/d): ").strip().upper()
                except (EOFError, KeyboardInterrupt):
                    print("\nOperacion cancelada por el usuario.")
                    opcion_export_texto = "D"
                    
                if opcion_export_texto == "":
                    print("Opcion vacia: seleccione a, b, c o d.")
                    continue
                    
                if opcion_export_texto == "D":
                    break
                    
                if opcion_export_texto not in ("A", "B", "C"):
                    print("Opcion invalida: seleccione a, b, c o d.")
                    continue
                    
                filas_export = generar_reporte_por_fecha_lista(fecha_consulta)
                if opcion_export_texto == "A":
                    exportar_reporte_csv(fecha_consulta, filas_export)
                elif opcion_export_texto == "B":
                    exportar_reporte_json(fecha_consulta, filas_export)
                elif opcion_export_texto == "C":
                    exportar_reporte_excel(fecha_consulta, filas_export)
                break
                
            break

    elif opcion == 5:
        print("\n" + "=" * 60)
        print("REGISTRAR NUEVO CLIENTE")
        print("=" * 60)
        cancelar_cliente = False

        while True:
            try:
                texto_nombre = input("\nIngrese el nombre del cliente o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_cliente = True
                break
                
            if texto_nombre.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_cliente = True
                break
                
            if texto_nombre == "":
                print("Nombre invalido: el campo 'Nombre' esta vacio.")
                continue
            if any(char.isdigit() for char in texto_nombre):
                print("Nombre invalido: no se aceptan digitos en el nombre.")
                continue
            if not texto_nombre.replace(" ", "").isalpha():
                print("Nombre invalido: solo letras y espacios son permitidos. Ejemplo: Maria")
                continue
            break
            
        if cancelar_cliente:
            continue

        while True:
            try:
                texto_apellidos = input("Ingrese los apellidos del cliente o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_cliente = True
                break
                
            if texto_apellidos.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_cliente = True
                break
                
            if texto_apellidos == "":
                print("Apellidos invalidos: el campo 'Apellidos' esta vacio.")
                continue
            if any(char.isdigit() for char in texto_apellidos):
                print("Apellidos invalidos: no se aceptan digitos en los apellidos.")
                continue
            if not texto_apellidos.replace(" ", "").isalpha():
                print("Apellidos invalidos: solo letras y espacios son permitidos. Ejemplo: Garcia Lopez")
                continue
            break
            
        if cancelar_cliente:
            continue

        # Insertar cliente en BD
        try:
            asegurar_tablas()
            with sqlite3.connect(DB_FILE) as conexion:
                cursor = conexion.cursor()
                cursor.execute("INSERT INTO clientes(nombre,apellidos) VALUES(?,?)", (texto_nombre, texto_apellidos))
                conexion.commit()
                cliente_id_bd = cursor.lastrowid
                cursor.close()
                
            cargar_estado_desde_bd()
            print(f"\nCliente registrado exitosamente con ID: {cliente_id_bd}")
            print(f"Nombre: {texto_nombre} {texto_apellidos}")
            
        except sqlite3.IntegrityError as err:
            print(f"Cliente no insertado en BD (error de integridad): {err}")
        except Exception as err:
            print(f"Error al insertar cliente en BD: {err}")

    elif opcion == 6:
        print("\n" + "=" * 60)
        print("REGISTRAR NUEVA SALA")
        print("=" * 60)
        cancelar_sala = False

        while True:
            try:
                texto_nombre_sala = input("\nIngrese el nombre de la sala o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_sala = True
                break
                
            if texto_nombre_sala.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_sala = True
                break
                
            if texto_nombre_sala == "":
                print("Nombre de sala invalido: campo vacio.")
                continue
            if any(char.isdigit() for char in texto_nombre_sala):
                print("Nombre de sala invalido: no se aceptan digitos en el nombre de sala.")
                continue
            if not all(car.isalpha() or car.isspace() for car in texto_nombre_sala):
                print("Nombre de sala invalido: solo letras y espacios permitidos. Ejemplo: Sala Ejecutiva")
                continue
            break
            
        if cancelar_sala:
            continue

        while True:
            try:
                texto_cupo = input("Ingrese el cupo de la sala (entero mayor a 0) o 'X' para cancelar: ").strip()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                cancelar_sala = True
                break
                
            if texto_cupo.upper() == "X":
                print("Operacion cancelada por el usuario.")
                cancelar_sala = True
                break
                
            if texto_cupo == "":
                print("Cupo invalido: campo vacio.")
                continue
            if any(c.isalpha() for c in texto_cupo):
                print("Cupo invalido: no se aceptan letras en el cupo.")
                continue
            if not texto_cupo.isdigit():
                print("Cupo invalido: formato no numerico. Ejemplo valido: 12")
                continue
                
            try:
                cupo_int = int(texto_cupo)
            except ValueError:
                print("Cupo invalido: no se pudo convertir a entero.")
                continue
                
            if cupo_int == 0:
                print("Cupo invalido: la sala no puede tener cupo 0.")
                continue
            if cupo_int < 0:
                print("Cupo invalido: no se aceptan numeros negativos.")
                continue
            break
            
        if cancelar_sala:
            continue

        # Insertar sala en BD
        try:
            asegurar_tablas()
            with sqlite3.connect(DB_FILE) as conexion:
                cursor = conexion.cursor()
                cursor.execute("INSERT INTO salas(nombre,cupo) VALUES(?,?)", (texto_nombre_sala, cupo_int))
                conexion.commit()
                sala_id_bd = cursor.lastrowid
                cursor.close()
                
            cargar_estado_desde_bd()
            print(f"\nSala registrada exitosamente con ID: {sala_id_bd}")
            print(f"Nombre: {texto_nombre_sala}")
            print(f"Cupo: {cupo_int} personas")
            
        except sqlite3.IntegrityError as err:
            print(f"Sala no insertada en BD (error de integridad): {err}")
        except Exception as err:
            print(f"Error al insertar sala en BD: {err}")

    elif opcion == 7:
        while True:
            try:
                respuesta_salir = input("\nEsta seguro que desea salir del programa? (S/N): ").strip().upper()
            except (EOFError, KeyboardInterrupt):
                print("\nOperacion cancelada por el usuario.")
                respuesta_salir = "N"
                
            if respuesta_salir == "":
                print("Entrada vacia: indique 'S' para salir o 'N' para cancelar.")
                continue
            if respuesta_salir not in ("S", "N"):
                print("Opcion invalida: solo 'S' o 'N'.")
                continue
            break
            
        if respuesta_salir == "S":
            print("\n" + "=" * 70)
            print("¡Gracias por usar el Sistema de Reservacion!".center(70))
            print("Saliendo del programa...".center(70))
            print("=" * 70)
            sys.exit()
        else:
            print("Continuando en el programa...")

    else:
        print("Opcion no valida. Intente de nuevo.")